
import openpyxl 

#load workbook 
wb=openpyxl.load_workbook("final.xlsx")

# Load sheet 
sh1=wb['Sheet1']

#get max number of rows in sheet
row=sh1.max_row

#get max number of columns in sheet
column=sh1.max_column


print(row,column)


data=[('Num','Name','Result'),(1,'Mukesh',90),(2,'Python',99),(3,'Java',95)]

#run for loop which will read all records from sheet one by one
#for i in range(1,row+5):
#    for j in range(1,column+1):
#        print(sh1.cell(i,j).value)
#run for loop and append the record one by one
for i in data:
    sh1.append(i)
#save the workbook 
wb.save("NewExcelDemo1.xlsx")