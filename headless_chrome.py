# Importing all necessary libraries
import time
import xlsxwriter
import openpyxl
import mysql.connector

from selenium import webdriver
from selenium.webdriver.support.select import Select
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options


user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36"

options = webdriver.ChromeOptions()
options.headless = True
#options.add_argument(f'user-agent={user_agent}')
#options.add_argument("--window-size=1920,1080")
#options.add_argument('--ignore-certificate-errors')
#options.add_argument('--allow-running-insecure-content')
#options.add_argument("--disable-extensions")
#options.add_argument("--proxy-server='direct://'")
#options.add_argument("--proxy-bypass-list=*")
#options.add_argument("--start-maximized")
#options.add_argument('--disable-gpu')
#options.add_argument('--disable-dev-shm-usage')
#options.add_argument('--no-sandbox')




# Defining the connection to the database

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="",
  database="scraper"
)
mycursor = mydb.cursor()




# Links required to fulfill the fetching
browser = webdriver.Chrome('C:\\Users\\marli\\Downloads\\chromedriver_win32\\chromedriver.exe',options=options)
base_link = 'https://www.kfzteile24.de'
braking_system_link = 'https://www.kfzteile24.de/ersatzteile-verschleissteile/bremsanlage?ktypnr='
pre = 'https://www.kfzteile24.de/ersatzteile-verschleissteile/bremsanlage/bremsbelaege?ktypnr='
#browser.get('https://www.kfzteile24.de/ersatzteile-verschleissteile/bremsanlage/bremsscheiben?ktypnr=1325')

def scroll_bottom(total_brake_disc):
    # Scroll down to the bottom of the page
    l=browser.find_element_by_xpath("//*[contains(text(), 'Copyright © 2021 kfzteile24.de - Alle Rechte vorbehalten')]")
    # action object creation to scroll
    a = ActionChains(browser)
    a.move_to_element(l).perform()
    if(total_brake_disc > 200):
      sleep(20)
    if(total_brake_disc > 150):
      sleep(15)
    if(total_brake_disc > 100):
      sleep(10)
    sleep(int(total_brake_disc/10))

def fetching(total_brake_disc):

    # Fetching the content from the web page
    names  =  browser.find_elements_by_class_name("art-name")
    price = browser.find_elements_by_class_name("preis")
    image = browser.find_elements_by_class_name('art-name')
    n_list=[]
    p_list=[]
    c_list=[]
    
    for nm in names:
        s=nm.text
        s=s.splitlines()
        if(len(s)>1):
            s=s[1]
            s= s.strip('Art.-Nr. ')
            n_list.append(s)
  
    for ps in price:
        ps = ps.text
        p_list.append(ps)
    
    for i in image:

        try:
            n=i.find_element_by_tag_name('img')
            n=n.get_attribute('alt')
            c_list.append(n)
        except:
            pass
    
    print("Total names: ",len(names))
    print("Names fetched: ",len(n_list))
    print("Price fetched: ",len(p_list))    
    print("Images fetched: ",len(c_list))
    
    if(len(n_list) == total_brake_disc and len(p_list) == total_brake_disc and len(c_list) == total_brake_disc and len(n_list) == len(p_list) and len(n_list) == len(c_list) ):
        return n_list,p_list,c_list
    else:
        print("Retrying:")
        return fetching(total_brake_disc)

def post_fetching(brand_name,model_name,type_name,n_list,p_list,c_list):

    # Making a list for appending in the excel file
    final_list = []
    #brand_type = [brand_name]*len((n_list))
    car_type=[type_name]*len(n_list)
    current_model_name = [model_name]*len(n_list)
    for i in range(len(n_list)):
            final_list.append([current_model_name[i], car_type[i],"brake_pads" ,n_list[i], p_list[i], c_list[i]])
    return(tuple(final_list))

'''
def excel_task(past_excel_name, excel_name, final_list):
    
    #load previous workbook 
    wb=openpyxl.load_workbook(past_excel_name)
    # Load sheet 
    sh1=wb['new']
    #get max number of rows in sheet
    row=sh1.max_row
    #get max number of columns in sheet
    column=sh1.max_column
    print("Row=",row,"\nColumn=",column)
    # Appeding the previous and making a new file 
    for f in final_list:
        sh1.append(f)
    wb.save(excel_name)
'''
def insert_into_database(final_data,current_brand):
    sql = "INSERT INTO "+current_brand+" (model,type,product_name,product_id,price,dealer) VALUES (%s, %s, %s, %s, %s, %s)"
    for i in range(len(final_data)):
        i=tuple(final_data[i])
        print(i)
        print(type(i))
        mycursor.execute(sql, i)
        mydb.commit()
    return

def breaking_system(id):
    #brake_disc_list = []
    url = braking_system_link + str(id)
    browser.get(url)
    categories = browser.find_element_by_class_name('categories')
    catgeory_data=categories.text
    category=catgeory_data.splitlines()
    print(category)
    brake_disc = "Bremsbeläge"
    if(brake_disc in category):
        brake_disc_index = category.index("Bremsbeläge")
        brake_disc_total = str(category[brake_disc_index+1])
        brake_disc_total = brake_disc_total.lstrip("(").rsplit(" ")
        brake_disc_total = int(brake_disc_total[0])
        #brake_disc_list.append(brake_disc_total)
        #print(type(brake_disc_total)," - ",brake_disc_total)
        return(brake_disc_total)
    else:
        return(0)

def select_brand(current_brand):

  browser.get(base_link)
  brand_select = Select(browser.find_element_by_class_name("brandSelector"))
  brand_select.select_by_visible_text(current_brand)
  print("Current Brand - ",current_brand)
  sleep(2)

def select_model(current_model):
  models = browser.find_element_by_class_name("modelSelector")
  Select_model = Select(models)
  Select_model.select_by_index(current_model)
  print("Current Model - ",current_model)

def reach_path():
    #start = 1
    brand_list_1 = ["AUDI", "BMW",'SEAT', 'SKODA', 'TOYOTA', 'VOLVO', 'VW', 'APOLLO', 'ARO', 'ARTEGA', 'ASIA MOTORS', 'ASTON MARTIN', 'AUDI', 'AUSTIN', 
    'AUSTIN_HEALEY', 'AUTO_UNION', 'AUTOBIANCHI', 'BARKAS', 'BEDFORD', 'BENTLEY', 'BERTONE', 'BITTER', 'BIZZARRINI', 'BMW', 'BORGWARD', 
    'BRILLIANCE', 'BRISTOL', 'BUGATTI', 'BUICK', 'CADILLAC', 'CALLAWAY', 'CATERHAM', 'CHEVROLET', 'CHRYSLER', 'CITROËN', 'CUPRA', 'DACIA', 
    'DAEWOO', 'DAF', 'DAIHATSU', 'DAIMLER', 'DE LOREAN', 'DE TOMASO', 'DFSK', 'DODGE', 'DONKERVOORT', 'DS', 'eGO', 'FERRARI', 'FIAT', 'FISKER', 
    'FORD', 'FORD USA', 'FSO', 'GAZ', 'GEELY', 'GENESIS', 'GERMAN E-CARS', 'GLAS', 'GMC', 'GOUPIL', 'GUMPERT', 'HONDA', 'HUMMER', 'HYUNDAI', 
    'INFINITI', 'INNOCENTI', 'IRMSCHER', 'ISDERA', 'ISO', 'ISORIVOLTA', 'ISUZU', 'IVECO', 'JAGUAR', 'JEEP', 'KIA', 'KOENIGSEGG', 'KTM', 'LADA', 
    'LAMBORGHINI', 'LANCIA', 'LAND_ROVER', 'LANDWIND', 'LDV', 'LEVC', 'LEXUS', 'LIGIER', 'LINCOLN', 'LLOYD', 'LOTUS', 'LTI', 'MAHINDRA', 
    'MAN', 'MARCOS', 'MASERATI', 'MAXUS', 'MAYBACH', 'MAZDA', 'MCLAREN', 'MEGA', 'MERCEDES-BENZ', 'METROCAB', 'MG', 'MIA_ELECTRIC', 'MICROCAR', 
    'MINELLI', 'MINI', 'MITSUBISHI', 'MITSUOKA', 'MORGAN', 'MORRIS', 'MOSKVICH', 'NISSAN', 'NSU', 'OLDSMOBILE', 'OPEL', 'PAGANI', 'PEUGEOT', 
    'PIAGGIO', 'PININFARINA', 'PLYMOUTH', 'POLARIS', 'POLESTAR', 'PONTIAC', 'PORSCHE', 'PROTON', 'PUCH', 'RENAULT', 'RENAULT TRUCKS', 
    'REVA', 'RILEY', 'ROLLS-ROYCE', 'ROVER', 'RUF', 'SAAB', 'SANTANA', 'SEAT', 'SHUANGHUAN', 'SKODA', 'SMART', 'SPYKER', 'SSANGYONG', 
    'STEYR', 'STREETSCOOTER', 'SUBARU', 'SUZUKI', 'TALBOT', 'TATA', 'TAZZARI', 'TESLA', 'THINK', 'TOYOTA', 'TRABANT', 'TRIUMPH', 'TVR', 'UAZ', 
    'VAUXHALL', 'VECTOR', 'VOLVO', 'VW', 'WARTBURG', 'WIESMANN']
    brand_list = ["ZASTAVA"]
    for brand in range(len(brand_list)):
        #print("----------------------------------------\n")
        #print("Model -> ",l)
        
        current_brand = brand_list[brand]
        create_table_query = "CREATE TABLE "+current_brand+ " (id INT AUTO_INCREMENT PRIMARY KEY, model VARCHAR(255), type VARCHAR(255), product_name VARCHAR(255), product_id VARCHAR(255), price VARCHAR(255), dealer VARCHAR(255))" 
        print(create_table_query)
        mycursor.execute(create_table_query)
        # Brand Selection
        '''
        brand_select = Select(browser.find_element_by_class_name("brandSelector"))
        current_brand = brand_list[brand]
        print("Current Brand - ",current_brand)
        brand_select.select_by_visible_text(current_brand)
        sleep(2)
        '''

        # Function calling for brand selection
        select_brand(current_brand)

        # All Model Selection and removing model 0
        model_value_list = []
        model_name_list = []
        models = browser.find_element_by_class_name("modelSelector")
        total_models = [x for x in models.find_elements_by_tag_name("option")]
        #print(total_models)
        for model in total_models:
            model_value_list.append(model.get_attribute("value"))
            model_name_list.append(model.text)
        
        sleep(2)
        model_value_list.pop(0)
        #model_name_list.pop(0)
        print("---- Total Model for",current_brand, " brand are:  ",model_name_list)
        print(model_value_list)
        #current_model = 9
        for current_model in range(len(model_name_list)):
            
            # Again brand selection
            select_brand(current_brand)
            sleep(1)
            try:
              print("in try")
              select_model(current_model)
            except:
              print("in except")
              select_brand(current_brand)
              return select_model(current_model)

            sleep(2)

            # All Type Selection and removing type 0
            type_selector = browser.find_element_by_class_name("typeSelector")
            types = [x for x in type_selector.find_elements_by_tag_name("option")]
            type_value=[]
            type_name=[]
            #print("options: ",options)
            for data in types:
                type_value.append(data.get_attribute("value"))
                type_name.append(data.text)
            type_value.pop(0)
            type_name.pop(0)

            audi_type_option = type_name
            audi_type_option_value = type_value


            current_model_name = model_name_list[current_model]
        
            print("Current Model-> ",current_model_name)
            print("Type Names-> ",type_name)
            print("Type values-> ",type_value)

            #print(type(current_model_name),current_model_name)
            

            
            for j in range(len(audi_type_option_value)):
                total_brake_disc = breaking_system(audi_type_option_value[j])
                if(total_brake_disc==0):
                    continue

                print("Car type option - ",audi_type_option[j])
                print("Car Type value - ",audi_type_option_value[j])
                print("Total data - ",total_brake_disc)
                audi_type= audi_type_option[j]
                
                post = audi_type_option_value[j]
                url = pre + str(post)
                browser.get(url)
                scroll_bottom(total_brake_disc)
                #excel_name = str(post)+'.xlsx'
                '''
                if(start==1):
                    n_list,p_list,c_list = fetching(total_brake_disc)
                    final_list = post_fetching(current_brand,current_model_name,audi_type,n_list,p_list,c_list)
                    excel_task("test.xlsx",excel_name,final_list)
                    start+=1
                    past_type = audi_type_option_value[j]
                    past_excel_name = str(past_type)+'.xlsx'
                
                else:
                '''    
                n_list,p_list,c_list = fetching(total_brake_disc)
                final_list = post_fetching(current_brand,current_model_name,audi_type,n_list,p_list,c_list)
                insert_into_database(final_list,current_brand)
                #excel_task(past_excel_name, excel_name, final_list)
                #past_type = audi_type_option_value[j]
                #past_excel_name = str(past_type)+'.xlsx'
               

reach_path()
browser.close()
'''
driver = webdriver.Chrome('C:\\Users\\marli\\Downloads\\chromedriver_win32\\chromedriver.exe')
options = Options()
options.headless = True
driver = webdriver.Chrome(options=options)
driver.get("https://www.youtube.com/")
element_text = driver.find_element_by_id("title").text
print(element_text)
'''