import time
import xlsxwriter
import openpyxl

from selenium import webdriver
from selenium.webdriver.support.select import Select
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains


browser = webdriver.Chrome('C:\\Users\\marlin.shah\\Downloads\\chromedriver_win32\\chromedriver')
base_link = 'https://www.kfzteile24.de'
braking_system_link = 'https://www.kfzteile24.de/ersatzteile-verschleissteile/bremsanlage?ktypnr='
pre = 'https://www.kfzteile24.de/ersatzteile-verschleissteile/bremsanlage/bremsscheiben?ktypnr='

#browser.get('https://www.kfzteile24.de/ersatzteile-verschleissteile/bremsanlage/bremsscheiben?ktypnr=1325')



#car_manufacture = ['AUDI', 'BMW', 'CITROËN', 'FIAT', 'FORD', 'HONDA', 'HYUNDAI', 'KIA', 'MAZDA', 'MERCEDES', 'BENZ', 'MITSUBISHI', 'NISSAN', 'OPEL', 'PEUGEOT', 'RENAULT', 'SEAT', 'SKODA', 'TOYOTA', 'VOLVO', 'VW', 'ABARTH', 'AC', 'AIXAM', 'ALFA', 'ROMEO', 'ALPINA', 'ALPINE', 'AMC', 'APOLLO', 'ARO', 'ARTEGA', 'ASIA', 'MOTORS', 'ASTON', 'MARTIN', 'AUDI', 'AUSTIN', 'AUSTIN', 'HEALEY', 'AUTO', 'UNION', 'AUTOBIANCHI', 'BARKAS', 'BEDFORD', 'BENTLEY', 'BERTONE', 'BITTER', 'BIZZARRINI', 'BMW', 'BORGWARD', 'BRILLIANCE', 'BRISTOL', 'BUGATTI', 'BUICK', 'CADILLAC', 'CALLAWAY', 'CATERHAM', 'CHEVROLET', 'CHRYSLER', 'CITROËN', 'CUPRA', 'DACIA', 'DAEWOO', 'DAF', 'DAIHATSU', 'DAIMLER', 'DE', 'LOREAN', 'DE', 'TOMASO', 'DFSK', 'DODGE', 'DONKERVOORT', 'DS', 'e.GO', 'FERRARI', 'FIAT', 'FISKER', 'FORD', 'FORD', 'USA', 'FSO', 'GAZ', 'GEELY', 'GENESIS', 'GERMAN', 'E-CARS', 'GLAS', 'GMC', 'GOUPIL', 'GUMPERT', 'HONDA', 'HUMMER', 'HYUNDAI', 'INFINITI', 'INNOCENTI', 'IRMSCHER', 'ISDERA', 'ISO', 'ISORIVOLTA', 'ISUZU', 'IVECO', 'JAGUAR', 'JEEP', 'KIA', 'KOENIGSEGG', 'KTM', 'LADA', 'LAMBORGHINI', 'LANCIA', 'LAND', 'ROVER', 'LANDWIND', '(JMC)', 'LDV', 'LEVC', 'LEXUS', 'LIGIER', 'LINCOLN', 'LLOYD', 'LOTUS', 'LTI', 'MAHINDRA', 'MAN', 'MARCOS', 'MASERATI', 'MAXUS', 'MAYBACH', 'MAZDA', 'MCLAREN', 'MEGA', 'MERCEDES-BENZ', 'METROCAB', 'MG', 'MIA', 'ELECTRIC', 'MICROCAR', 'MINELLI', 'MINI', 'MITSUBISHI', 'MITSUOKA', 'MORGAN', 'MORRIS', 'MOSKVICH', 'NISSAN', 'NSU', 'OLDSMOBILE', 'OPEL', 'PAGANI', 'PEUGEOT', 'PIAGGIO', 'PININFARINA', 'PLYMOUTH', 'POLARIS', 'POLESTAR', 'PONTIAC', 'PORSCHE', 'PROTON', 'PUCH', 'RENAULT', 'RENAULT', 'TRUCKS', 'REVA', '(MAHINDRA)', 'RILEY', 'ROLLS-ROYCE', 'ROVER', 'RUF', 'SAAB', 'SANTANA', 'SEAT', 'SHUANGHUAN', 'SKODA', 'SMART', 'SPYKER', 'SSANGYONG', 'STEYR', 'STREETSCOOTER', 'SUBARU', 'SUZUKI', 'TALBOT', 'TATA', 'TAZZARI', 'TESLA', 'THINK', 'TOYOTA', 'TRABANT', 'TRIUMPH', 'TVR', 'UAZ', 'VAUXHALL', 'VECTOR', 'VOLVO', 'VW', 'WARTBURG', 'WIESMANN', 'ZASTAVA', 'ZAZ', 'ZUENDAPP']

# audi_model_option = ['100 (43, C2), 06/76 bis 09/82', '100 (44, 44Q, C3), 08/82 bis 07/91', '100 (4A2, C4), 12/90 bis 07/94', '100 (C1), 11/68 bis 07/76', '100 Avant (43, C2), 07/77 bis 02/83', '100 Avant (44, 44Q, C3), 08/82 bis 11/90', '100 Avant (4A5, C4), 12/90 bis 11/94', '100 Coupe (C1), 01/70 bis 12/76', '200 (43), 10/79 bis 09/82', '200 (44, 44Q), 06/83 bis 12/91', '200 Avant (44, 44Q), 09/83 bis 12/91', '50 (86), 08/74 bis 07/78', '60, 09/68 bis 08/72', '60 Variant, 09/68 bis 08/72', '75, 09/68 bis 08/72', '75 Variant, 09/68 bis 08/72', '80, 09/66 bis 08/68', '80 (80, 82, B1), 05/72 bis 07/78', '80 (81, 85, B2), 08/78 bis 03/87', '80 (89, 89Q, 8A, B3), 06/86 bis 10/91', '80 (8C2, B4), 09/91 bis 07/95', '80 Avant (8C5, B4), 09/91 bis 01/96', '80 Variant, 09/66 bis 08/68', '90 (81, 85, B2), 08/84 bis 03/87', '90 (89, 89Q, 8A, B3), 04/87 bis 09/91', 'A1 (8X1, 8XK), 05/10 bis 10/18', 'A1 CITY CARVER (GBH), ab 07/19', 'A1 Sportback (8XA, 8XF), 09/11 bis 10/18', 'A1 Sportback (GBA), ab 07/18', 'A2 (8Z0), 02/00 bis 08/05', 'A3 (8L1), 09/96 bis 09/06', 'A3 (8P1), 05/03 bis 12/13', 'A3 (8V1, 8VK), ab 04/12', 'A3 Cabriolet (8P7), 04/08 bis 05/13', 'A3 Cabriolet (8V7, 8VE), ab 10/13', 'A3 Limousine (8VS, 8VM), ab 05/13', 'A3 Limousine (8YS), ab 04/20', 'A3 Sportback (8PA), 09/04 bis 12/15', 'A3 Sportback (8VA, 8VF), ab 09/12', 'A3 Sportback (8YA), ab 11/19', 'A4 (8D2, B5), 11/94 bis 12/01', 'A4 (8E2, B6), 11/00 bis 12/05', 'A4 (8EC, B7), 11/04 bis 06/08', 'A4 (8K2, B8), 08/07 bis 12/15', 'A4 (8W2, 8WC, B9), ab 05/15', 'A4 Allroad (8KH, B8), 04/09 bis 05/16', 'A4 Allroad (8WH, 8WJ, B9), ab 01/16', 'A4 Avant (8D5, B5), 11/94 bis 12/02', 'A4 Avant (8E5, B6), 02/00 bis 12/05', 'A4 Avant (8ED, B7), 11/04 bis 06/08', 'A4 Avant (8K5, B8), 11/07 bis 12/15', 'A4 Avant (8W5, 8WD, B9), ab 08/15', 'A4 Cabriolet (8H7, B6, 8HE, B7), 01/02 bis 12/09', 'A5 (8T3), 06/07 bis 01/17', 'A5 (F53, F5P), ab 06/16', 'A5 Cabriolet (8F7), 02/09 bis 01/17', 'A5 Cabriolet (F57, F5E), ab 11/16', 'A5 Sportback (8TA), 07/07 bis 01/17', 'A5 Sportback (F5A, F5F), ab 06/16', 'A6 (4A2, C4), 06/94 bis 10/97', 'A6 (4A2, C8), ab 02/18', 'A6 (4B2, C5), 01/97 bis 08/05', 'A6 (4F2, C6), 04/04 bis 08/11', 'A6 (4G2, 4GC, C7), 11/10 bis 09/18', 'A6 Allroad (4AH, C8), ab 11/18', 'A6 Allroad (4FH, C6), 03/06 bis 08/11', 'A6 Allroad (4GH, 4GJ, C7), 01/12 bis 09/18', 'A6 Avant (4A5, C4), 06/94 bis 12/97', 'A6 Avant (4A5, C8), ab 05/18', 'A6 Avant (4B5, C5), 11/97 bis 01/05', 'A6 Avant (4F5, C6), 11/04 bis 08/11', 'A6 Avant (4G5, 4GD, C7), 05/11 bis 09/18', 'A7 Sportback (4GA, 4GF), 07/10 bis 05/18', 'A7 Sportback (4KA), ab 10/17', 'A8 (4D2, 4D8), 03/94 bis 12/05', 'A8 (4E2, 4E8), 10/02 bis 12/10', 'A8 (4H2, 4H8, 4HC, 4HL), 11/09 bis 01/18', 'A8 (4N2, 4N8), ab 06/17', 'ALLROAD (4BH, C5), 05/00 bis 08/05', 'CABRIOLET (8G7, B4), 05/91 bis 08/00', 'COUPE (81, 85), 07/80 bis 10/88', 'COUPE (89, 8B), 10/88 bis 12/96', 'E-TRON (GEN_), ab 09/18', 'E-TRON Sportback (GEA), ab 09/19', 'F103, 08/65 bis 08/68', 'F103 Variant, 04/66 bis 08/68', 'Q2 (GAB), ab 06/16', 'Q3 (8UB, 8UG), 06/11 bis 10/18', 'Q3 (F3B), ab 07/18', 'Q3 Sportback (F3N), ab 06/19', 'Q5 (8RB), 11/08 bis 12/17', 'Q5 (FYB), ab 05/16', 'Q7 (4LB), 03/06 bis 01/16', 'Q7 (4MB, 4MG), ab 01/15', 'Q8 (4MN), ab 02/18', 'QUATTRO (85), 07/80 bis 07/91', 'R8 (422, 423), 04/07 bis 07/15', 'R8 (4S3, 4SP), ab 07/15', 'R8 Spyder (427, 429), 02/10 bis 07/15', 'R8 Spyder (4S9, 4SR), ab 05/16', 'SUPER 90, 09/66 bis 11/71', 'TT (8J3), 07/06 bis 01/15', 'TT (8N3), 10/98 bis 06/06', 'TT (FV3, FVP), ab 07/14', 'TT Roadster (8J9), 03/07 bis 06/14', 'TT Roadster (8N9), 10/99 bis 06/06', 'TT Roadster (FV9, FVR), ab 11/14', 'V8 (441, 442, 4C2, 4C8), 10/88 bis 02/94']
# audi_model_option_value = ['10', '13', '19', '53', '25', '30', '36', '48', '172', '176', '1940', '433', '1500', '1501', '1502', '4553', '8554', '14', '26', '31', '1', '6', '8555', '78', '84', '8604', '40147', '9730', '38933', '4382', '1557', '4955', '10253', '6724', '11611', '11282', '40796', '5143', '10459', '40531', '253', '4731', '5376', '6418', '14695', '8031', '36378', '3485', '4797', '5377', '6988', '14696', '4840', '6243', '36928', '8030', '37293', '8159', '37116', '258', '38843', '1891', '5110', '9154', '39591', '5546', '10161', '3468', '38844', '3395', '5381', '9627', '8986', '38471', '265', '4935', '8627', '38219', '4632', '438', '194', '202', '39213', '40160', '8552', '8553', '36543', '9731', '38997', '40155', '7534', '37161', '5461', '14159', '38836', '3910', '6244', '14514', '8262', '36765', '4554', '5544', '3851', '12912', '5608', '4089', '13274', '485']

# audi_type_option 3 = ['2.0, 100 PS, 74 kW, 12/90 bis 07/94', '2.0 E, 115 PS, 85 kW, 12/90 bis 07/94', '2.0 E 16V, 140 PS, 103 kW, 01/92 bis 07/94', '2.0 E quattro, 115 PS, 85 kW, 12/90 bis 07/92', '2.3 E, 133 PS, 98 kW, 12/90 bis 07/94', '2.3 E quattro, 133 PS, 98 kW, 12/90 bis 07/94', '2.4 D, 82 PS, 60 kW, 12/90 bis 07/94', '2.5 TDI, 115 PS, 85 kW, 12/90 bis 07/94', '2.6, 150 PS, 110 kW, 03/92 bis 07/94', '2.6 quattro, 150 PS, 110 kW, 04/92 bis 07/94', '2.8 E, 174 PS, 128 kW, 12/90 bis 07/94', '2.8 E quattro, 174 PS, 128 kW, 12/90 bis 07/94', 'S4 Turbo quattro, 230 PS, 169 kW, 08/91 bis 07/94', 'S4 V8 quattro, 280 PS, 206 kW, 10/92 bis 07/94']
#audi_type_option_value 3 = ['1381', '1383', '12481', '4996', '1384', '1388', '1379', '1380', '1385', '1389', '1386', '1390', '1387', '4670']

#audi_type_option 7  = ['2.0 E, 100 PS, 74 kW, 09/91 bis 07/94', '2.0 E, 115 PS, 85 kW, 09/91 bis 07/94', '2.0 E 16V, 140 PS, 103 kW, 07/92 bis 07/94', '2.0 E quattro, 115 PS, 85 kW, 12/90 bis 07/92', '2.3 E, 133 PS, 98 kW, 09/91 bis 07/94', '2.3 E quattro, 133 PS, 98 kW, 09/91 bis 07/94', '2.4 D, 82 PS, 60 kW, 12/90 bis 07/94', '2.5 TDI, 115 PS, 85 kW, 12/90 bis 07/94', '2.6, 150 PS, 110 kW, 03/92 bis 07/94', '2.6 quattro, 150 PS, 110 kW, 07/92 bis 07/94', '2.8 E, 174 PS, 128 kW, 09/91 bis 07/94', '2.8 E quattro, 174 PS, 128 kW, 09/91 bis 07/94', 'S4 Turbo quattro, 230 PS, 169 kW, 09/91 bis 07/94', 'S4 V8 quattro, 280 PS, 206 kW, 10/92 bis 07/94']
#audi_type_option_value 7 = ['1447', '1448', '20106', '4997', '1452', '1468', '5006', '1444', '1450', '1469', '1453', '1471', '1449', '4669']



def scroll_bottom(total_brake_disc):
    # Scroll down to the bottom of the page
    l=browser.find_element_by_xpath("//*[contains(text(), 'Copyright © 2021 kfzteile24.de - Alle Rechte vorbehalten')]")
    # action object creation to scroll
    a = ActionChains(browser)
    a.move_to_element(l).perform()
    sleep(int(total_brake_disc/10))

def fetching(total_brake_disc):

    # Fetching the content from the web page
    names  =  browser.find_elements_by_class_name("art-name")
    price = browser.find_elements_by_class_name("preis")
    image = browser.find_elements_by_class_name('art-name')
    n_list=[]
    p_list=[]
    c_list=[]
    
    for nm in names:
        s=nm.text
        s=s.splitlines()
        if(len(s)>1):
            s=s[1]
            s= s.strip('Art.-Nr. ')
            n_list.append(s)
  
    for ps in price:
        ps = ps.text
        p_list.append(ps)
    
    for i in image:

        try:
            n=i.find_element_by_tag_name('img')
            n=n.get_attribute('alt')
            c_list.append(n)
        except:
            pass
    
    print("Total names: ",len(names))
    print("Names fetched: ",len(n_list))
    print("Price fetched: ",len(p_list))    
    print("Images fetched: ",len(c_list))
    
    if(len(n_list) == total_brake_disc and len(p_list) == total_brake_disc and len(c_list) == total_brake_disc and len(n_list) == len(p_list) and len(n_list) == len(c_list) ):
        return n_list,p_list,c_list
    else:
        print("Retrying:")
        return fetching(total_brake_disc)

def post_fetching(brand_name,model_name,type_name,n_list,p_list,c_list):

    # Making a list for appending in the excel file
    final_list = []
    brand_type = [brand_name]*len((n_list))
    car_type=[type_name]*len(n_list)
    current_model_name = [model_name]*len(n_list)
    for i in range(len(n_list)):
            final_list.append(['Bremsscheiben',brand_type[i], current_model_name[i], car_type[i], n_list[i], p_list[i], c_list[i]])
    return(final_list)

def excel_task(past_excel_name, excel_name, final_list):
    
    #load previous workbook 
    wb=openpyxl.load_workbook(past_excel_name)

    # Load sheet 
    sh1=wb['new']

    #get max number of rows in sheet
    row=sh1.max_row

    #get max number of columns in sheet
    column=sh1.max_column

    print("Row=",row,"\nColumn=",column)
    # Appeding the previous and making a new file 
    for f in final_list:
        sh1.append(f)
    wb.save(excel_name)
  
def breaking_system(id):
    #brake_disc_list = []
    url = braking_system_link + str(id)
    browser.get(url)
    categories = browser.find_element_by_class_name('categories')
    catgeory_data=categories.text
    category=catgeory_data.splitlines()
    print(category)
    brake_disc = "Bremsscheiben"
    if(brake_disc in category):
        brake_disc_index = category.index("Bremsscheiben")
        brake_disc_total = str(category[brake_disc_index+1])
        brake_disc_total = brake_disc_total.lstrip("(").rsplit(" ")
        brake_disc_total = int(brake_disc_total[0])
        #brake_disc_list.append(brake_disc_total)
        #print(type(brake_disc_total)," - ",brake_disc_total)
        return(brake_disc_total)
    else:
        return(0)

def reach_path():
    start = 1
    brand_list = ["FIAT"]
    for brand in range(len(brand_list)):
        #print("----------------------------------------\n")
        #print("Model -> ",l)
        browser.get(base_link)

        # Brand Selection

        brand_select = Select(browser.find_element_by_class_name("brandSelector"))
        current_brand = brand_list[brand]
        print("Current Brand - ",current_brand)
        brand_select.select_by_visible_text(current_brand)
        sleep(2)

        # Model Selection
        model_value_list = []
        model_name_list = []
        models = browser.find_element_by_class_name("modelSelector")
        total_models = [x for x in models.find_elements_by_tag_name("option")]
        #print(total_models)
        for model in total_models:
            model_value_list.append(model.get_attribute("value"))
            model_name_list.append(model.text)
        print("Total Model = ",model_name_list)
        sleep(2)
        model_value_list.pop(0)
        #current_model = 19
        for current_model in range(56,len(model_value_list)):
            current_model_name = model_name_list[current_model]
            browser.get(base_link)
            brand_select = Select(browser.find_element_by_class_name("brandSelector"))
            current_brand = brand_list[brand]
            print("Current Brand - ",current_brand)
            sleep(1)
            brand_select.select_by_visible_text(current_brand)
            models = browser.find_element_by_class_name("modelSelector")
            Select_model = Select(models)
            Select_model.select_by_index(current_model)
            sleep(2)

            # Type Selection
            type_selector = browser.find_element_by_class_name("typeSelector")
            types = [x for x in type_selector.find_elements_by_tag_name("option")]
            type_value=[]
            type_name=[]
            #print("options: ",options)
            for data in types:
                type_value.append(data.get_attribute("value"))
                type_name.append(data.text)
            type_value.pop(0)
            type_name.pop(0)
            print("Type Names-> ",type_name)
            print("Type values-> ",type_value)
            audi_type_option = type_name
            audi_type_option_value = type_value
            for j in range(len(audi_type_option_value)):
                total_brake_disc = breaking_system(audi_type_option_value[j])
                if(total_brake_disc==0):
                    continue
                print("Current Model Name - ",current_model_name)   
                print("Car type option - ",audi_type_option[j])
                print("Car Type value - ",audi_type_option_value[j])
                print("Total data - ",total_brake_disc)
                sleep(int(total_brake_disc/10))
                audi_type= audi_type_option[j]
                
                post = audi_type_option_value[j]
                url = pre + str(post)
                browser.get(url)
                scroll_bottom(total_brake_disc)
                excel_name = str(post)+'.xlsx'
                
                if(start==1):

                    n_list,p_list,c_list = fetching(total_brake_disc)
                    final_list = post_fetching(current_brand,current_model_name,audi_type,n_list,p_list,c_list)
                    excel_task("15592.xlsx",excel_name,final_list)
                    start+=1
                    past_type = audi_type_option_value[j]
                    past_excel_name = str(past_type)+'.xlsx'
                
                else:
                    
                    n_list,p_list,c_list = fetching(total_brake_disc)
                    final_list = post_fetching(current_brand,current_model_name,audi_type,n_list,p_list,c_list)
                    excel_task(past_excel_name, excel_name, final_list)
                    past_type = audi_type_option_value[j]
                    past_excel_name = str(past_type)+'.xlsx'
               
              

reach_path()
browser.close()
