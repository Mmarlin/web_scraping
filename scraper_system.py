import time
import xlrd
import openpyxl
import mysql.connector

from selenium import webdriver
from selenium.webdriver.support.select import Select
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains

# Defining the connection to the database

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="mshah",
  database="test"
)
mycursor = mydb.cursor()

options = webdriver.ChromeOptions()
options.headless = False
browser = webdriver.Chrome('C:\\Users\\marlin.shah\\Downloads\\chromedriver_win32\\chromedriver',options=options)

base_link = 'https://www.kfzteile24.de'

system = 'braking_system'

# Open the workbook and define the worksheet
book = xlrd.open_workbook("C:\\Users\\marlin.shah\\Desktop\\scraper_masterxls.xls")
sheet = book.sheet_by_name(system)

mycursor.execute("SELECT model,type FROM " + system)
db_data = mycursor.fetchall()
for i in range(len(db_data)):
    db_data[i] = list(db_data[i])
last_model_name = db_data[-1][0]
# Create the INSERT INTO sql query
query = "INSERT INTO "+ system + " (ktype, make, model, type, part_category, part_name, cat1, cat2, cat3, cat4, part_no, price, brand) VALUES (%s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s)"



count = 0
def scroll_bottom():
    # Scroll down to the bottom of the page
    l=browser.find_element_by_xpath("//*[contains(text(), 'Copyright © 2021 kfzteile24.de - Alle Rechte vorbehalten')]")
    # action object creation to scroll
    a = ActionChains(browser)
    a.move_to_element(l).perform()
    sleep(5)
   
def fetching(count):
    try:
        browser.find_element_by_xpath("//*[contains(text(), '0 Teile passend für:')]")
        return None,None,None
    except:
        
        # Fetching the content from the web page
        names  =  browser.find_elements_by_class_name("art-name")
        price = browser.find_elements_by_class_name("preis")
        image = browser.find_elements_by_class_name('art-name')
        n_list=[]
        p_list=[]
        c_list=[]
        
        for nm in names:
            s=nm.text
            s=s.splitlines()
            if(len(s)>1):
                s=s[1]
                s= s.strip('Art.-Nr. ')
                n_list.append(s)
    
        for ps in price:
            ps = ps.text
            p_list.append(ps)
        
        for i in image:

            try:
                n=i.find_element_by_tag_name('img')
                n=n.get_attribute('alt')
                c_list.append(n)
            except:
                pass
        
        print("Total names: ",len(names))
        print("Names fetched: ",len(n_list))
        print("Price fetched: ",len(p_list))    
        print("Images fetched: ",len(c_list))
        
        if(count > 5):
            print(n_list,p_list,c_list)
            if(len(n_list) == len(p_list) and len(n_list) == len(c_list) ):   
                count = 0
                print("reached")
                return n_list,p_list,c_list
            
        else:
            print("Retrying:",count)
            count +=1
          
            sleep(int(len(names)/10))
        
            return fetching(count)

def post_fetching(ktype, make, model, model_type, part_category, part_name, cat1, cat2, cat3, cat4, part_no, price, brand):

    # Making a list for appending in the excel file
    final_list = []
    #brand_type = [brand_name]*len((n_list))
    #car_type=[type_name]*len(n_list)
    #current_model_name = [model_name]*len(n_list)
    
    for i in range(len(part_no)):
        final_list.append([ktype, make, model, model_type, part_category, part_name, cat1, cat2, cat3, cat4, part_no[i], price[i], brand[i]])
    return(tuple(final_list))

def insert_into_database(final_data):
    
    for i in range(len(final_data)):
        i=tuple(final_data[i])
        print(i)
        print(type(i))
        mycursor.execute(query, i)
        mydb.commit()
    return

def select_brand(current_brand):

  browser.get(base_link)
  brand_select = Select(browser.find_element_by_class_name("brandSelector"))
  brand_select.select_by_visible_text(current_brand)
  print("Current Brand - ",current_brand)
  sleep(2)

def select_model(current_model):

  models = browser.find_element_by_class_name("modelSelector")
  Select_model = Select(models)
  Select_model.select_by_index(current_model)

def reach_path():
  
    brand_list = ["AUDI"]
    for brand in range(len(brand_list)):
 
        browser.get(base_link)
        current_brand = brand_list[brand]
    
        # Brand Selection
        # Function calling for brand selection
        select_brand(current_brand)

        # Model Selection
        model_value_list = []
        model_name_list = []
        models = browser.find_element_by_class_name("modelSelector")
        total_models = [x for x in models.find_elements_by_tag_name("option")]
        #print(total_models)
        for model in total_models:
            model_value_list.append(model.get_attribute("value"))
            model_name_list.append(model.text)
        print("Total Model for",current_brand, " brand are:  ",model_name_list)
        sleep(2)
        model_value_list.pop(0)
        last_model_id = model_name_list.index(last_model_name)
        print("Last model id: ",last_model_id)
        for current_model in range(last_model_id,109):
            current_model_name = model_name_list[current_model]
            print(type(current_model_name),current_model_name,"Id is: ",current_model)
            # Again brand selection
            select_brand(current_brand)
            sleep(1)
            try:
              print("in try")
              select_model(current_model)
              print("Current Model - ",current_model_name)
            except:
              print("in except")
              select_brand(current_brand)
              print("Current Model - ",current_model_name)
              return select_model(current_model)


            sleep(2)

            # Type Selection
            type_selector = browser.find_element_by_class_name("typeSelector")
            types = [x for x in type_selector.find_elements_by_tag_name("option")]
           
            type_value=[]
            type_name=[]
            #print("options: ",options)
            for data in types:
                type_value.append(data.get_attribute("value"))
                type_name.append(data.text)
              
            type_value.pop(0)
            type_name.pop(0)

            print("Type Names-> ",type_name)
            print("Type values-> ",type_value)
            audi_type_option = type_name
            audi_type_option_value = type_value
            for j in range(len(audi_type_option_value)):
                check = [current_model_name,audi_type_option[j]]
                print(check)
                if(check not in db_data):
                    for r in range(2, sheet.nrows):
                        pre= sheet.cell(r,1).value
                        part_category = sheet.cell(r,2).value
                        part_name = sheet.cell(r,3).value
                        cat1 = sheet.cell(r,4).value
                        cat2 = sheet.cell(r,5).value
                        cat3 = sheet.cell(r,6).value
                        cat4 = sheet.cell(r,7).value

                        print("Car type option - ",audi_type_option[j])
                        print("Car Type value - ",audi_type_option_value[j])
                        print("Position: ",part_category,part_name)
                        audi_type= audi_type_option[j]

                        post = audi_type_option_value[j]
                        url = pre + '?ktypnr=' + str(post)
                        browser.get(url)
                        scroll_bottom()
                        
                        
                            
                        n_list,p_list,c_list = fetching(count)
                        if(n_list == None):
                            continue
                        final_data = post_fetching(post, current_brand, current_model_name, audi_type, part_category, part_name, cat1, cat2, cat3, cat4, n_list, p_list, c_list)
                        insert_into_database(final_data)
                    

reach_path()

browser.close()



print("done!")

