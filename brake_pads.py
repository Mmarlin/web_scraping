import time
import xlsxwriter
import openpyxl

from selenium import webdriver
from selenium.webdriver.support.select import Select
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains


browser = webdriver.Chrome('C:\\Users\\marlin.shah\\Downloads\\chromedriver_win32\\chromedriver')
base_link = 'https://www.kfzteile24.de'
braking_system_link = 'https://www.kfzteile24.de/ersatzteile-verschleissteile/bremsanlage?ktypnr='
pre = 'https://www.kfzteile24.de/ersatzteile-verschleissteile/bremsanlage/bremsscheiben?ktypnr='

brake_pad_link = 'https://www.kfzteile24.de/ersatzteile-verschleissteile/bremsanlage/bremsbelaege'

#browser.get('https://www.kfzteile24.de/ersatzteile-verschleissteile/bremsanlage/bremsscheiben?ktypnr=1325')
def scroll_bottom(total_brake_disc):
    # Scroll down to the bottom of the page
    l=browser.find_element_by_xpath("//*[contains(text(), 'Copyright © 2021 kfzteile24.de - Alle Rechte vorbehalten')]")
    # action object creation to scroll
    a = ActionChains(browser)
    a.move_to_element(l).perform()
    if(total_brake_disc > 200):
      sleep(20)
    if(total_brake_disc > 150):
      sleep(15)
    if(total_brake_disc > 100):
      sleep(10)
    sleep(int(total_brake_disc/10))

def fetching(total_brake_disc):

    # Fetching the content from the web page
    names  =  browser.find_elements_by_class_name("art-name")
    price = browser.find_elements_by_class_name("preis")
    image = browser.find_elements_by_class_name('art-name')
    n_list=[]
    p_list=[]
    c_list=[]
    
    for nm in names:
        s=nm.text
        s=s.splitlines()
        if(len(s)>1):
            s=s[1]
            s= s.strip('Art.-Nr. ')
            n_list.append(s)
  
    for ps in price:
        ps = ps.text
        p_list.append(ps)
    
    for i in image:

        try:
            n=i.find_element_by_tag_name('img')
            n=n.get_attribute('alt')
            c_list.append(n)
        except:
            pass
    
    print("Total names: ",len(names))
    print("Names fetched: ",len(n_list))
    print("Price fetched: ",len(p_list))    
    print("Images fetched: ",len(c_list))
    
    if(len(n_list) == total_brake_disc and len(p_list) == total_brake_disc and len(c_list) == total_brake_disc and len(n_list) == len(p_list) and len(n_list) == len(c_list) ):
        return n_list,p_list,c_list
    else:
        print("Retrying:")
        return fetching(total_brake_disc)

def post_fetching(brand_name,model_name,type_name,n_list,p_list,c_list):

    # Making a list for appending in the excel file
    final_list = []
    brand_type = [brand_name]*len((n_list))
    car_type=[type_name]*len(n_list)
    current_model_name = [model_name]*len(n_list)
    for i in range(len(n_list)):
            final_list.append(['Bremsscheiben',brand_type[i], current_model_name[i], car_type[i], n_list[i], p_list[i], c_list[i]])
    return(final_list)

def excel_task(past_excel_name, excel_name, final_list):
    
    #load previous workbook 
    wb=openpyxl.load_workbook(past_excel_name)

    # Load sheet 
    sh1=wb['new']

    #get max number of rows in sheet
    row=sh1.max_row

    #get max number of columns in sheet
    column=sh1.max_column

    print("Row=",row,"\nColumn=",column)
    # Appeding the previous and making a new file 
    for f in final_list:
        sh1.append(f)
    wb.save(excel_name)
  
def breaking_system(id):
    #brake_disc_list = []
    url = braking_system_link + str(id)
    browser.get(url)
    categories = browser.find_element_by_class_name('categories')
    catgeory_data=categories.text
    category=catgeory_data.splitlines()
    print(category)
    brake_disc = "Bremsscheiben"
    if(brake_disc in category):
        brake_disc_index = category.index("Bremsscheiben")
        brake_disc_total = str(category[brake_disc_index+1])
        brake_disc_total = brake_disc_total.lstrip("(").rsplit(" ")
        brake_disc_total = int(brake_disc_total[0])
        #brake_disc_list.append(brake_disc_total)
        #print(type(brake_disc_total)," - ",brake_disc_total)
        return(brake_disc_total)
    else:
        return(0)

def select_brand(current_brand):

  browser.get(brake_pad_link)
  brand_select = Select(browser.find_element_by_class_name("brandSelector"))
  brand_select.select_by_visible_text(current_brand)
  print("Current Brand - ",current_brand)
  sleep(2)


def total_brands():
    brand_name_list = []
    brand_value_list = []
    browser.get(base_link)
    brand_select = Select(browser.find_element_by_class_name("brandSelector"))
    total_brands = [x for x in brand_select.find_elements_by_tag_name("option")]
    for brand in total_brands:
        brand_value_list.append(brand.get_attribute("value"))
        brand_name_list.append(brand.text)
    
    return brand_name_list,brand_value_list


def select_model(current_model):
  models = browser.find_element_by_class_name("modelSelector")
  Select_model = Select(models)
  Select_model.select_by_index(current_model)

def reach_path():
    browser.get(base_link)
    brand_name_list, brand_value_list = total_brands()
    browser.get(brake_pads)
    start = 1
    
    for current_brand in brand_value_list:

        #browser.get(base_link)
        #current_brand = brand_list[brand] 
        # Brand Selection
        '''
        brand_select = Select(browser.find_element_by_class_name("brandSelector"))
        current_brand = brand_list[brand]
        print("Current Brand - ",current_brand)
        brand_select.select_by_visible_text(current_brand)
        sleep(2)
        '''

        # Function calling for brand selection
        select_brand(current_brand)

        # Model Selection
        model_value_list = []
        model_name_list = []
        models = browser.find_element_by_class_name("modelSelector")
        total_models = [x for x in models.find_elements_by_tag_name("option")]
        #print(total_models)
        for model in total_models:
            model_value_list.append(model.get_attribute("value"))
            model_name_list.append(model.text)
        print("---- Total Model for",current_brand, " brand are:  ",model_name_list)
        sleep(2)
        model_value_list.pop(0)
        #current_model = 9
        for current_model in range(len(model_value_list)):
            current_model_name = model_name_list[current_model]
            print(type(current_model_name),current_model_name)
            # Again brand selection
            select_brand(current_brand)
            sleep(1)
            try:
              print("in try")
              select_model(current_model)
              print("Current Model - ",current_model_name)
            except:
              print("in except")
              select_brand(current_brand)
              print("Current Model - ",current_model_name)
              return select_model(current_model)


            sleep(2)

            # Type Selection
            type_selector = browser.find_element_by_class_name("typeSelector")
            types = [x for x in type_selector.find_elements_by_tag_name("option")]
            type_value=[]
            type_name=[]
            #print("options: ",options)
            for data in types:
                type_value.append(data.get_attribute("value"))
                type_name.append(data.text)
            type_value.pop(0)
            type_name.pop(0)
            print("Type Names-> ",type_name)
            print("Type values-> ",type_value)
            audi_type_option = type_name
            audi_type_option_value = type_value
            for j in range(len(audi_type_option_value)):
                total_brake_disc = breaking_system(audi_type_option_value[j])
                if(total_brake_disc==0):
                    continue

                print("Car type option - ",audi_type_option[j])
                print("Car Type value - ",audi_type_option_value[j])
                print("Total data - ",total_brake_disc)
                audi_type= audi_type_option[j]
                
                post = audi_type_option_value[j]
                url = pre + str(post)
                browser.get(url)
                scroll_bottom(total_brake_disc)
                excel_name = str(post)+'.xlsx'
                
                if(start==1):

                    n_list,p_list,c_list = fetching(total_brake_disc)
                    final_list = post_fetching(current_brand,current_model_name,audi_type,n_list,p_list,c_list)
                    excel_task("test.xlsx",excel_name,final_list)
                    start+=1
                    past_type = audi_type_option_value[j]
                    past_excel_name = str(past_type)+'.xlsx'
                
                else:
                    
                    n_list,p_list,c_list = fetching(total_brake_disc)
                    final_list = post_fetching(current_brand,current_model_name,audi_type,n_list,p_list,c_list)
                    excel_task(past_excel_name, excel_name, final_list)
                    past_type = audi_type_option_value[j]
                    past_excel_name = str(past_type)+'.xlsx'
               
                

reach_path()
browser.close()